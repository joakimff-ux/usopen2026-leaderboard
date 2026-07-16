"""Import roster and scores from a fantasy golf Excel workbook."""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
TEAM_HEADER_ROW = 3
TEAM_COLUMN_START = 7
PLAYER_NAME_COLUMN = 1
ROUND_COLUMNS = {2: 1, 3: 2, 4: 3, 5: 4}
ROSTER_MARKERS = {"X", "x"}


@dataclass
class ParsedPlayer:
    name: str
    tier: int
    team_names: list[str] = field(default_factory=list)
    scores: dict[int, int] = field(default_factory=dict)


@dataclass
class ParsedWorkbook:
    teams: list[str]
    players: list[ParsedPlayer]
    team_roster_counts: dict[str, int]


def _normalize_name(value: Any) -> str:
    return str(value).strip()


def _is_tier_label(value: str) -> bool:
    return value.lower().startswith("tier ")


def _parse_tier(value: str) -> int:
    return int(value.split()[-1])


def _is_roster_mark(value: Any) -> bool:
    if value is None:
        return False
    return str(value).strip() in ROSTER_MARKERS


def _parse_score(value: Any) -> int | None:
    if value is None or str(value).strip() == "":
        return None
    return int(value)


def parse_workbook(path: str | Path) -> ParsedWorkbook:
    workbook_path = Path(path)
    worksheet = load_workbook(workbook_path, data_only=True)["Ark1"]

    first_tier_row = next(
        (
            row
            for row in range(1, worksheet.max_row + 1)
            if _is_tier_label(_normalize_name(worksheet.cell(row, PLAYER_NAME_COLUMN).value or ""))
        ),
        None,
    )

    # The Open workbook has participant names on the first tier row from C.
    # Keep support for the legacy result layout (row 3 from G, scores in B:E).
    if first_tier_row is not None and worksheet.cell(first_tier_row, 3).value:
        team_header_row = first_tier_row
        team_column_start = 3
        round_columns: dict[int, int] = {}
    else:
        team_header_row = TEAM_HEADER_ROW
        team_column_start = TEAM_COLUMN_START
        round_columns = ROUND_COLUMNS

    team_columns: list[tuple[int, str]] = []
    for column in range(team_column_start, worksheet.max_column + 1):
        header = worksheet.cell(team_header_row, column).value
        if header is None:
            if team_columns:
                break
            continue
        team_columns.append((column, _normalize_name(header)))
    teams = [team_name for _, team_name in team_columns]

    if not teams:
        raise ValueError("No team headers found in row 3 starting at column G.")

    players: list[ParsedPlayer] = []
    current_tier: int | None = None

    for row in range(first_tier_row or (TEAM_HEADER_ROW + 1), worksheet.max_row + 1):
        cell_value = worksheet.cell(row, PLAYER_NAME_COLUMN).value
        if cell_value is None:
            continue

        label = _normalize_name(cell_value)
        if not label:
            continue

        if _is_tier_label(label):
            current_tier = _parse_tier(label)
            continue

        if current_tier is None:
            raise ValueError(f"Player '{label}' on row {row} appears before any tier section.")

        roster_teams = [
            team_name
            for column, team_name in team_columns
            if _is_roster_mark(worksheet.cell(row, column).value)
        ]

        scores: dict[int, int] = {}
        for column, round_num in round_columns.items():
            parsed_score = _parse_score(worksheet.cell(row, column).value)
            if parsed_score is not None:
                scores[round_num] = parsed_score

        players.append(
            ParsedPlayer(
                name=label,
                tier=current_tier,
                team_names=roster_teams,
                scores=scores,
            )
        )

    if not players:
        raise ValueError("No players found in workbook.")

    team_roster_counts = {team_name: 0 for team_name in teams}
    for player in players:
        for team_name in player.team_names:
            team_roster_counts[team_name] += 1

    invalid_teams = [
        f"{team_name} ({count})"
        for team_name, count in team_roster_counts.items()
        if count != 7
    ]
    if invalid_teams:
        raise ValueError(
            "Each team must have exactly 7 golfers. Invalid teams: " + ", ".join(invalid_teams)
        )

    return ParsedWorkbook(teams=teams, players=players, team_roster_counts=team_roster_counts)


def import_workbook(client: Any, path: str | Path, replace_existing: bool = True) -> dict[str, Any]:
    from lib.db import get_active_tournament, reset_tournament_data, tournament_display_title

    parsed = parse_workbook(path)
    tournament = get_active_tournament(client)
    if tournament is None:
        raise RuntimeError("No active tournament found in tournaments table.")

    if replace_existing:
        reset_tournament_data(client, tournament["id"])

    team_rows = [
        {"tournament_id": tournament["id"], "name": team_name}
        for team_name in parsed.teams
    ]
    inserted_teams = client.table("teams").insert(team_rows).execute().data or []
    team_id_by_name = {team["name"]: team["id"] for team in inserted_teams}

    player_rows = [
        {"tournament_id": tournament["id"], "name": player.name, "tier": player.tier}
        for player in parsed.players
    ]
    inserted_players = client.table("players").insert(player_rows).execute().data or []
    player_id_by_name = {player["name"]: player["id"] for player in inserted_players}

    roster_rows = []
    for player in parsed.players:
        player_id = player_id_by_name[player.name]
        for team_name in player.team_names:
            roster_rows.append({"team_id": team_id_by_name[team_name], "player_id": player_id})

    if roster_rows:
        client.table("team_players").insert(roster_rows).execute()

    score_rows = []
    for player in parsed.players:
        player_id = player_id_by_name[player.name]
        for round_num, strokes in player.scores.items():
            score_rows.append(
                {
                    "player_id": player_id,
                    "round": round_num,
                    "strokes": strokes,
                    "source": "EXCEL",
                    "is_official": True,
                }
            )

    if score_rows:
        client.table("scores").upsert(score_rows, on_conflict="player_id,round").execute()

    return {
        "tournament_id": tournament["id"],
        "tournament_name": tournament_display_title(tournament),
        "teams_imported": len(parsed.teams),
        "players_imported": len(parsed.players),
        "roster_links": len(roster_rows),
        "scores_imported": len(score_rows),
        "team_roster_counts": parsed.team_roster_counts,
    }
